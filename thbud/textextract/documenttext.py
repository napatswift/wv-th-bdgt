from typing import List, Optional, Callable, Tuple, Union
import fitz
from .text import WordText, PageText, LineText
from ..tableparser import (
    has_table
)
import numpy as np
import openpyxl


def group_text_by_line(text_list: List[WordText], threshold=0.01):
    """
    Group text by line

    Args:
      text_list (List[WordText]): list of WordText objects representing text in pdf file.
      threshold (float, optional): threshold to group text. Defaults to 0.01.

    Returns:
      List[List[WordText]]: a list of lines, where each line is a list of WordText objects.
    """
    # Sort text by y0, x0.
    text_list.sort(key=lambda x: (x.y0, x.x0))
    line_list = []
    line = []
    prev_y0 = None
    for word in text_list:
        # If previous y0 is None, set it to y0.
        if prev_y0 is None:
            prev_y0 = word.y0

        # If current y0 - previous y0 > threshold,
        # add current line to line list and start a new line.
        if word.y0 - prev_y0 > threshold:
            line_list.append(line)
            line = []

        # Add current text to line.
        line.append(word)

        # Update previous y0.
        prev_y0 = word.y0

    # If line is not empty, add it to line list.
    if line:
        line_list.append(line)

    # Sort by x0
    for line in line_list:
        line.sort(key=lambda x: x.x0)

    return line_list


def is_image_page(page: fitz.Page) -> bool:
    """
    Check if a page is an image page.

    Args:
        page (fitz.Page): The page to check.

    Returns:
        bool: True if the page is an image page, False otherwise.

    Raises:
        TypeError: If page is not a fitz.Page object.
    """
    if not isinstance(page, fitz.Page):
        raise TypeError('page must be a fitz.Page object.')

    theshold = 0.8
    page_size = page.rect.width * page.rect.height

    # Get images in the page.
    # images: List[(xref, smask, width, height, bpc, colorspace, alt. colorspace, name, filter, referencer)]
    images = page.get_images()
    if not images:
        return False

    image_size = 0
    for image in images:
        image_size += image[2] * image[3]

    # If the size of the images is greater than theshold of the page size,
    # then it's an image page.
    return image_size / page_size > theshold


def is_rect_inside_page(rect, page_width, page_height):
    return rect.x0 >= 0 and rect.x1 <= page_width and rect.y0 >= 0 and rect.y1 <= page_height


def page_contains_table(page: fitz.Page) -> bool:
    pix = page.get_pixmap()
    image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
        pix.h, pix.w, pix.n)
    return has_table(image)


class DocumentText:
    def __init__(
        self,
        filepath: str,
        words_loader: Optional[Callable[[
            fitz.Page], List[Tuple[float, float, float, float, str]]]] = None,
        lazy: bool = False,
    ) -> 'DocumentText':
        self.filepath = filepath
        self.page_label_to_index = dict()  # str as key
        self.lazy = lazy
        self.doc = None
        self.pages = []
        if words_loader is None:
            words_loader = self.defualt_words_loader
        self.words_loader = words_loader
        self._read_pdf_file()

    def defualt_words_loader(
            self,
            page: fitz.Page
    ) -> List[Tuple[float, float, float, float, str]]:
        word_tuples = page.get_text_words()
        words = []
        for word in word_tuples:
            x0, y0, x1, y1, text = word[:5]
            words.append((x0, y0, x1, y1, text))

        return words

    def _read_pdf_file(self,) -> List['PageText']:
        try:
            doc: fitz.Document = fitz.open(self.filepath)
        except fitz.FileNotFoundError as e:
            raise FileNotFoundError(
                'File not found: {}'.format(self.filepath)) from e

        self.doc = doc

        for pidx, page in enumerate(doc.pages()):
            self.page_label_to_index[page.get_label()] = pidx
            self.pages.append(None)
            if self.lazy:
                continue
            self._load_page(pidx)

    def _load_page(self, page_index: int) -> 'PageText':
        for pidx, page in enumerate(self.doc.pages()):
            if pidx != page_index or self.pages[pidx] is not None:
                continue

            word_tuples = self.words_loader(page)
            page_width = page.rect.width
            page_height = page.rect.height

            words = []
            for word in word_tuples:
                x0, y0, x1, y1, text = word
                x0 = x0 / page_width
                x1 = x1 / page_width
                y0 = y0 / page_height
                y1 = y1 / page_height
                words.append(WordText(x0, y0, x1, y1, text))
            grouped_lines = group_text_by_line(words, 0.01)
            list_of_line: List['LineText'] = []

            for lidx, words in enumerate(grouped_lines):
                if words:
                    list_of_line.append(LineText(words, pidx, lidx))

            pagetext = PageText(lines=list_of_line,
                                page_index=pidx,
                                width=page.rect.width,
                                height=page.rect.height,
                                is_image=is_image_page(page),
                                document=self,)

            for line in pagetext.lines:
                line.page = pagetext

            pagetext.contains_table = page_contains_table(page)
            self.pages[pidx] = pagetext

    def get_page(self: 'DocumentText', page_index: int) -> 'PageText':
        if page_index < 0 or page_index >= len(self.pages):
            raise IndexError('page_index must be in range [0, {})'.format(
                len(self.pages)))
        if self.pages[page_index] is None:
            self._load_page(page_index)
        return self.pages[page_index]

    def get_lines_in_page(
            self: 'DocumentText',
            start: Optional[int] = None,
            end: Optional[int] = None
    ) -> List['LineText']:
        if start is None:
            start = 0
        elif isinstance(start, str):
            start = self.page_label_to_index.get(start)
            if start is None:
                raise IndexError(
                    'page {} not found in the doc {}'.format(repr(start), self.filepath))
        if end is None:
            end = len(self.pages)
        elif isinstance(end, str):
            end = self.page_label_to_index.get(end)
            if end is None:
                raise IndexError(
                    'page {} not found in the doc {}'.format(repr(end), self.filepath))

        if start < 0 or start >= len(self.pages):
            raise IndexError('start must be in range [0, {})'.format(
                len(self.pages)))

        if end < 0 or end > len(self.pages):
            raise IndexError('end must be in range [0, {})'.format(
                len(self.pages)))

        lines = []
        for idx in range(start, end):
            page = self.get_page(idx)
            lines.extend(page.lines)

        return lines


class XLSXDocumentText:
    def __init__(self, filepath: str) -> None:
        self.filepath = filepath
        self.pages = []
        self.doc = None
        self.sheet_name_to_index = dict()
        self._read_xlsx_file()

    def get_page(self, page_index: Union[int, str]) -> 'PageText':
        if isinstance(page_index, str):
            page_index = self.sheet_name_to_index.get(page_index)
            if page_index is None:
                raise IndexError(
                    'sheet {} not found in the doc {}'.format(
                        repr(page_index), self.filepath
                    )
                )
        if page_index < 0 or page_index >= len(self.pages):
            raise IndexError('page_index must be in range [0, {})'.format(
                len(self.pages)))
        return self.pages[page_index]

    def _get_cell_value(self, cell) -> str:
        if not cell.value:
            return ''

        if cell.data_type == 'n':
            return f'{cell.value:,}'
        if cell.data_type == 's':
            return cell.value

        return str(cell.value)

    def _page_contains_table(self, ws) -> bool:
        rows = ws.iter_rows()
        cell_with_border_count = 0
        cell_with_border_threshold = 3

        def at_least_n_borders(cell, n):
            return sum([
                1 for border in [
                    cell.border.left,
                    cell.border.right,
                    cell.border.top,
                    cell.border.bottom
                ] if border.style
            ]) >= n

        for row in rows:
            for cell in row:
                if at_least_n_borders(cell, 2):
                    cell_with_border_count += 1
                    if cell_with_border_count >= cell_with_border_threshold:
                        return True
        return False

    def _read_xlsx_file(self) -> None:
        wb = openpyxl.load_workbook(self.filepath)
        for sheet_index, sheet in enumerate(wb.sheetnames):
            ws = wb[sheet]
            hidden_columns = {
                column_index: v.hidden
                for k, v in ws.column_dimensions.items()
                for column_index in range(v.min, v.max + 1)
            }
            lines: List['LineText'] = []
            for row in ws.iter_rows():
                words = []
                row_cumulative_indent = 0
                for cell in row:
                    if (
                        cell.value
                        and not hidden_columns.get(cell.col_idx, False)
                    ):
                        row_cumulative_indent += cell.alignment.indent
                        words.append(WordText(
                            x0=cell.column + row_cumulative_indent,
                            y0=cell.row,
                            x1=cell.column + 1 + row_cumulative_indent,
                            y1=cell.row,
                            text=self._get_cell_value(cell)
                        ))

                line = LineText(words, page_index=sheet_index,
                                line_index=cell.row)
                lines.append(line)
            self.sheet_name_to_index[sheet] = sheet_index
            page = PageText(lines, sheet_index, 1, 1, False, self)

            for line in page.lines:
                line.page = page
            
            page.contains_table = self._page_contains_table(ws)

            self.pages.append(page)

    def get_lines_in_page(
            self,
            start: Optional[Union[int, str]] = None,
            end: Optional[Union[int, str]] = None
    ) -> List['LineText']:
        if start is None:
            start = 0
        elif isinstance(start, str):
            start = self.sheet_name_to_index.get(start)
            if start is None:
                raise IndexError(
                    'sheet {} not found in the doc {}'.format(
                        repr(start), self.filepath
                    )
                )
        if end is None:
            end = len(self.pages)
        elif isinstance(end, str):
            end = self.sheet_name_to_index.get(end)
            if end is None:
                raise IndexError(
                    'sheet {} not found in the doc {}'.format(
                        repr(end), self.filepath
                    )
                )

        if start < 0 or start >= len(self.pages):
            raise IndexError('start must be in range [0, {})'.format(
                len(self.pages)))

        if end < 0 or end > len(self.pages):
            raise IndexError('end must be in range [0, {})'.format(
                len(self.pages)))

        lines = []
        for idx in range(start, end):
            page = self.get_page(idx)
            lines.extend(page.lines)

        return lines
